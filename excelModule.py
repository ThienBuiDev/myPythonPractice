import openpyxl
# row and column start at 1, index starts at 0
workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook['Sheet1']
#print(workbook.get_sheet_names) # returns a list of sheet names
cell = sheet['A1']
#sheet.cell(row = 1,column = 2) get cell value using numbers for row and column can be used for for loop
print(sheet['B3'].value)

"""---Editing excel"""
wb = openpyxl.Workbook() #Open blank work book
# New workbook just have one sheet 'Sheet' and empty
wb.get_sheet_names()
wb['Sheet']['A1'] = 'Hello'

wb.save('myNewExcel.xlsx')
sheet2 = wb.create_sheet(index = 0) #add a new sheet
sheet2.title = 'new sheet name' #change sheet name